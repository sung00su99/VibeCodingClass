"""
엑셀 파일을 SQLite DB로 임포트하는 스크립트.
사용법: python import_excel.py <엑셀파일경로>

엑셀 컬럼 순서 (헤더명으로 매핑):
  DEPT, NAME, TITLE, PHONE, EMAIL, CLASSYN, REMARKS
"""
import sys
import sqlite3
import os

try:
    import openpyxl
except ImportError:
    print("openpyxl 패키지가 필요합니다: pip install openpyxl")
    sys.exit(1)

from database import get_connection, init_db

COLUMN_MAP = {
    "dept":              "DEPT",
    "name":              "NAME",
    "title":             "TITLE",
    "phone":             "PHONE",
    "email":             "EMAIL",
    "classyn":           "CLASSYN",
    "remarks":           "REMARKS",
    # 한국어 헤더
    "부서":              "DEPT",
    "이름":              "NAME",
    "직함":              "TITLE",
    "휴대폰":            "PHONE",
    "전자 메일 주소":    "EMAIL",
    "참석유무(y/n)":     "CLASSYN",
    "비고":              "REMARKS",
}


def import_excel(filepath: str):
    if not os.path.exists(filepath):
        print(f"파일을 찾을 수 없습니다: {filepath}")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

    init_db()
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM Vibe_Coding_Class_Info")
    seq = cursor.fetchone()[0] + 1

    inserted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        record = {}
        for idx, col_key in enumerate(headers):
            if col_key in COLUMN_MAP and idx < len(row):
                db_col = COLUMN_MAP[col_key]
                record[db_col] = str(row[idx]).strip() if row[idx] is not None else ""

        if not record.get("NAME") or not record.get("EMAIL"):
            print(f"  건너뜀 (NAME/EMAIL 필수): {record}")
            continue

        cursor.execute(
            """
            INSERT INTO Vibe_Coding_Class_Info
              (SEQNO, DEPT, NAME, TITLE, PHONE, EMAIL, CLASSYN, REMARKS)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                f"{seq:04d}",
                record.get("DEPT", ""),
                record.get("NAME", ""),
                record.get("TITLE", ""),
                record.get("PHONE", ""),
                record.get("EMAIL", ""),
                record.get("CLASSYN", ""),
                record.get("REMARKS", ""),
            ),
        )
        seq += 1
        inserted += 1

    conn.commit()
    conn.close()
    print(f"임포트 완료: {inserted}건 추가")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python import_excel.py <엑셀파일경로>")
        sys.exit(1)
    import_excel(sys.argv[1])
