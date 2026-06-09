"""
Excel → Supabase 직접 마이그레이션 스크립트
1. 기존 vibe_coding_class_info 테이블 전체 삭제 (dept != '기타' 포함 전체)
2. auto01.xlsx 읽어서 삽입

실행:
  export SUPABASE_SERVICE_KEY=<service_role key>
  python3 backend/excel_to_supabase.py excel/auto01.xlsx
"""

import sys
import json
import os
import urllib.request
import urllib.error

try:
    import openpyxl
except ImportError:
    print("openpyxl 필요: /tmp/venv2/bin/python3 로 실행하거나 pip install openpyxl")
    sys.exit(1)

SUPABASE_URL = "https://ktqkjvdzqxdkicvmlzni.supabase.co"
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")
TABLE = "vibe_coding_class_info"

HEADER_MAP = {
    "부서": "dept",
    "이름": "name",
    "직함": "title",
    "휴대폰": "phone",
    "전자 메일 주소": "email",
    "참석유무": "classyn",
    "참석유무(y/n)": "classyn",
    "비고": "remarks",
    # 영문 헤더도 지원
    "dept": "dept",
    "name": "name",
    "title": "title",
    "phone": "phone",
    "email": "email",
    "classyn": "classyn",
    "remarks": "remarks",
}


def supabase_request(method, path, body=None, params=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    if params:
        from urllib.parse import urlencode
        url += "?" + urlencode(params)
    headers = {
        "Content-Type": "application/json",
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Prefer": "return=minimal",
    }
    data = json.dumps(body).encode("utf-8") if body is not None else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            return resp.status, resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode("utf-8")


def delete_all():
    print(f"[1] {TABLE} 전체 삭제 중...")
    status, body = supabase_request("DELETE", TABLE, params={"id": "gte.0"})
    if status in (200, 204):
        print("    삭제 완료")
    else:
        print(f"    삭제 실패: HTTP {status}\n{body}")
        sys.exit(1)


def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    raw_headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    headers = [HEADER_MAP.get(h.lower(), HEADER_MAP.get(h, None)) for h in raw_headers]

    rows = []
    for seq_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if all(v is None for v in row):
            continue
        record = {"seqno": f"{seq_idx:04d}"}
        for idx, col in enumerate(headers):
            if col and idx < len(row):
                val = row[idx]
                record[col] = str(val).strip() if val is not None else ""
        if not record.get("name") or not record.get("email"):
            print(f"  건너뜀 (name/email 없음): {record}")
            continue
        rows.append(record)
    return rows


def insert_rows(rows):
    print(f"[2] {len(rows)}건 삽입 중...")
    status, body = supabase_request("POST", TABLE, body=rows)
    if status in (200, 201):
        print(f"    삽입 완료: {len(rows)}건")
    else:
        print(f"    삽입 실패: HTTP {status}\n{body}")
        sys.exit(1)


def main():
    if not SUPABASE_SERVICE_KEY:
        print("오류: SUPABASE_SERVICE_KEY 환경변수를 설정하세요.")
        print("  export SUPABASE_SERVICE_KEY=<service_role key>")
        sys.exit(1)

    filepath = sys.argv[1] if len(sys.argv) > 1 else "excel/auto01.xlsx"
    if not os.path.exists(filepath):
        print(f"파일 없음: {filepath}")
        sys.exit(1)

    print(f"파일: {filepath}")
    rows = read_excel(filepath)
    print(f"읽은 데이터: {len(rows)}건")
    if rows:
        print(f"  첫 행: {rows[0]}")

    delete_all()
    insert_rows(rows)
    print("완료.")


if __name__ == "__main__":
    main()
